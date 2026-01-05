from argparse import ArgumentParser
from hashlib import sha256
import json
from openpyxl import load_workbook, Workbook
from os import path
from typing import Dict, List

class Transaction:
    def __init__(self, date, uid, source, description, amount):
        self.date = date
        self.uid = uid
        self.source = source
        self.description = description
        self.amount = amount

    def __str__(self):
        return f"{self.date},{self.uid},{self.source},{self.description},{self.amount},"

    def to_list(self):
        return [self.date, self.uid, self.source, self.description, self.amount]

def parse_config(config_file: str) -> Dict:
    with open(config_file, 'r') as f:
        data = json.load(f)
    return data

def parse_imported_transactions(input_file: str, config: Dict) -> List[Transaction]:
    workbook = load_workbook(input_file)
    sheet = workbook.active

    index = {}
    for i in range(1, sheet.max_column + 1):
        index[sheet.cell(1, i).value] = i

    transactions = []
    for i in range(2, sheet.max_row + 1):
        transaction_string = ""
        for j in range(1, sheet.max_column + 1):
            transaction_string += str(sheet.cell(i, j).value)

        uid = sha256(transaction_string.encode()).hexdigest()

        amount = None
        if config["single_amount_column"]:
            amount = sheet.cell(i, index[config["amount_column"]]).value
            if config["negate_amount"]:
                amount *= -1
        else:
            debits = sheet.cell(i, index[config["debits_column"]]).value
            if config["negate_debits"]:
                debits *= -1

            credits = sheet.cell(i, index[config["credits_column"]]).value
            if config["negate_credits"]:
                credits *= -1

            amount = debits + credits

        transaction = Transaction(sheet.cell(i, index[config["date_column"]]).value,
                                  uid,
                                  config["name"],
                                  sheet.cell(i, index[config["description_column"]]).value,
                                  amount)

        transactions.append(transaction)

    return transactions

def update_transaction_database(output_file: str, transactions: List[Transaction]):
    workbook = None
    sheet = None
    if path.exists(output_file):
        workbook = load_workbook(output_file)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "UID", "Source", "Description", "Amount"])

    existing_uids = set([ sheet.cell(row, column=2).value for row in range(2, sheet.max_row + 1) ])
    new_uids = set([ t.uid for t in transactions ]) - existing_uids

    for transaction in transactions:
        if transaction.uid in new_uids:
            sheet.append(transaction.to_list())

    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, column=1).number_format = 'm/d/yyyy'

    workbook.save(output_file)

if __name__ == "__main__":
    parser = ArgumentParser(description='Parse transactions from different financial accounts into a common format.')
    parser.add_argument('input', help='Path to input file (xlsx). Export from your financial account.')
    parser.add_argument('--config', '-c', help='Path to config file (JSON). Describes input file format.')
    parser.add_argument('--output', '-o', default='transactions.xlsx', help='Path to output file (xlsx). Creates file or appends imported transactions to existing file (ignores duplicate transactions).')
    args = parser.parse_args()

    if not args.config:
        args.config = f"configs/{args.input.split('.')[0]}.json"

    config = parse_config(args.config)
    transactions = parse_imported_transactions(args.input, config)
    update_transaction_database(args.output, transactions)
