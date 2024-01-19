from argparse import ArgumentParser
from hashlib import sha256
import json
from openpyxl import load_workbook
from typing import Dict, List

class Transaction:
    def __init__(self, date, uid, source, description, amount):
        self.date = date
        self.uid = uid
        self.source = source
        self.description = description
        self.amount = amount

    def __str__(self):
        return f"{self.date},{self.uid},{self.source},{self.description},{self.amount},"

def parse_config(config_file: str) -> Dict:
    with open(config_file, 'r') as f:
        data = json.load(f)
    return data

def parse_imported_transactions(input_file: str, config: Dict) -> List[Transaction]:
    workbook = load_workbook(input_file)
    sheet = workbook.active

    index = {}
    for i in range(1, sheet.max_column + 1):
        index[sheet.cell(1, i).value] = i

    transactions = []
    for i in range(2, sheet.max_row + 1):
        transaction_string = ""
        for j in range(1, sheet.max_column + 1):
            transaction_string += str(sheet.cell(i, j).value)

        uid = sha256(transaction_string.encode()).hexdigest()

        transaction = Transaction(sheet.cell(i, index[config["date_column"]]).value,
                                  uid,
                                  config["name"],
                                  sheet.cell(i, index[config["description_column"]]).value,
                                  sheet.cell(i, index[config["amount_column"]]).value)

        transactions.append(transaction)


    return transactions

if __name__ == "__main__":
    parser = ArgumentParser(description='Parse transactions from different financial accounts into a common format.')
    parser.add_argument('input', help='Path to input file (xlsx). Export from your financial account.')
    parser.add_argument('config', help='Path to config file (JSON). Describes input file format.')
    parser.add_argument('output', help='Path to output file (xlsx). Creates file or appends imported transactions to existing file (ignores duplicate transactions).')
    args = parser.parse_args()

    config = parse_config(args.config)
    imported_transactions = parse_imported_transactions(args.input, config)
    for transaction in imported_transactions:
        print(str(transaction))

    # Add imported transactions to the transaction database (check transaction dates and hashes to avoid adding duplicates)

    # Make transaction parsing configurable for different inputs (decision log entry for JSON vs YAML)
    # Add instructions and examples to the README
    # Add simple testing
    # Add open source license
